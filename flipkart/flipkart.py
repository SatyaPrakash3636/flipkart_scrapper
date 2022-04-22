from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from flipkart import constant as const
from openpyxl import Workbook, load_workbook
from pathlib import Path
import time


class Flipkart(webdriver.Chrome):

    def __init__(self, teardown=False):
        self.service = Service(ChromeDriverManager().install())
        self.teardown = teardown
        super().__init__(service=self.service)
        self.implicitly_wait(15)
        self.maximize_window()

    def landing_page(self):
        self.get(const.BASE_URL)

    def login_page(self):
        login_button = self.find_element(By.CLASS_NAME, 'fGVtXS')
        login_button.click()
        user_name = self.find_element(
            By.CSS_SELECTOR, 'input[placeholder="Enter your username OR Phone Number OR Email"]'
        )
        user_name.send_keys(const.LOGIN_USER)
        time.sleep(3)
        next_button = self.find_element(
            By.XPATH, '//*[@id="app"]/div[3]/section/section/div/div[2]/button/span'
        )
        next_button.click()
        enter_pass = self.find_element(
            By.XPATH, '//*[@id="app"]/div[3]/section/section/div/div[1]/form/div[2]/div/input'
        )
        enter_pass.send_keys(const.LOGIN_PASS)
        login_button = self.find_element(
            By.XPATH, '//*[@id="app"]/div[3]/section/section/div/div[2]/button/span'
        )
        login_button.click()
        time.sleep(5)

    def select_seller(self):
        seller = self.find_element(
            By.XPATH, '//*[@id="blinx-wrapper-2"]/div/ul/li[1]/label/div/div[2]/div[1]/div[1]'
        )
        seller.click()
        access = self.find_element(
            By.XPATH, '//*[@id="blinx-wrapper-0"]/div[2]/div[3]/div/div[3]/button'
        )
        access.click()
        time.sleep(5)

    def write_data(self, wb, ws, header, fetched_data, my_file):
        for row in ws.iter_rows(min_row=1, max_col=9, max_row=1):
            for (cell, cell_data) in zip(row, header):
                cell.value = cell_data
        ws.insert_rows(2)
        for row in ws.iter_rows(min_row=2, max_col=9, max_row=2):
            for (cell, cell_data) in zip(row, fetched_data):
                cell.value = cell_data
        wb.save(rf'F:\Python_learning\selenium\flipkart\{my_file}')

    def get_data(self):
        seller = self.find_element(By.XPATH, '//*[@id="app-container"]/div/div[1]/div[2]/ul/li[5]/div/div/a/div/div[1]').text
        sales_period = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[1]/div[1]').text
        gross_sale = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[1]/div[2]/span').text
        unit_sold_element = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[2]/div[1]/div').text
        unit_sold = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[2]/div[2]/span').text
        dispatched_units_element = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[3]/div[1]/div[1]').text
        dispatched_units = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[3]/div[2]/span').text
        dispatched_sales_element = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[4]/div[1]/div[1]').text
        dispatched_sales = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[2]/div/div[4]/div[2]/span').text
        cancellations_element = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[2]/div[2]/span[1]/span[1]').text
        cancelled_amount = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[2]/div[3]/div[1]/div[1]/span').text
        cancelled_units = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[2]/div[3]/div[1]/div[2]/span').text
        returns_element = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[3]/div[2]/span[1]/span[1]').text
        returns_amount = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[3]/div[3]/div[1]/div[1]/span').text
        returns_unit = self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[2]/div[3]/div[1]/div/div[2]/div[3]/div[3]/div[1]/div[2]').text

        header = ["Sales Period", "Gross Sale", "Units Sold", "Dispatched Units", "Dispatched Sales", "Cancelled Amount", "Cancelled Units", "Returns Amount", "Returns Units"]
        fetched_data = [sales_period, gross_sale, int(unit_sold), int(dispatched_units), dispatched_sales, cancelled_amount, cancelled_units, returns_amount, returns_unit]

        print(f"Seller : {seller}")
        print(f"Sales Period : {sales_period}")
        print(f"Gross Sale : {gross_sale}")
        print(f"{unit_sold_element} : {unit_sold}")
        print(f"{dispatched_units_element} : {dispatched_units}")
        print(f"{dispatched_sales_element} : {dispatched_sales}")
        print(f"{cancellations_element}:")
        print(f"\tCancelled Amount : {cancelled_amount}")
        print(f"\tCancelled Units : {cancelled_units}")
        print(f"{returns_element}:")
        print(f"\tReturns Amount: {returns_amount}")
        print(f"\tReturns Units: {returns_unit}")
        print("#################################################################################")

        my_file = Path(r'F:\Python_learning\selenium\flipkart\earn_more.xlsx')
        if my_file.exists():
            wb = load_workbook(my_file)
            ws = wb[f"{seller}"]
            self.write_data(wb, ws, header, fetched_data, my_file)
        else:
            wb = Workbook()
            ws = wb.create_sheet(title=f"{seller}")
            self.write_data(wb, ws, header, fetched_data, my_file)

    def earn_more(self):
        try:
            skip = self.find_element(
                By.XPATH, '//*[@id="react-joyride:0"]/div/div/div[1]/div[2]/button[1]'
            )
            skip.click()
        except:
            pass
        wait = WebDriverWait(self, 10)
        actions = ActionChains(self)
        growth_menu = self.find_element(By.CSS_SELECTOR, "li[id=Growth]")
        actions.move_to_element(growth_menu).perform()
        wait.until(EC.visibility_of_element_located(
            (By.XPATH, '//*[@id="Growth"]/ul/li[4]/a'))).click()
        # click on month
        self.find_element(By.XPATH, '//*[@id="sub-app-container"]/section/section[1]/section/div[1]/div/div/div[3]/div').click()
        time.sleep(2)
        self.get_data()

        time.sleep(2)
        # for i in range(1, 4):
        #     self.find_element(By.XPATH,
        #                       '//*[@id="sub-app-container"]/section/section[1]/section/div[1]/div/div/div[3]/div/div/div/span').click()
        #     date_menu = self.find_element(By.XPATH,
        #                       '//*[@id="sub-app-container"]/section/section[1]/section/div[1]/div/div/div[3]/div/div/div/span')
        #     actions.move_to_element(date_menu).click()
        #     wait.until(EC.visibility_of_element_located((By.XPATH,
        #                       f'//div[@id="sub-app-container"]/section[1]/section[1]/section[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]/div[{i}]'))).click()
        #     time.sleep(2)
        #     self.get_data()
        #     time.sleep(2)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()


